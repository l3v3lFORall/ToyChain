
import utils.plugin as _up
from utils.common import pInfo


class customPlugin(_up.Plugin):
    """
    infoGather 工具ENScanGO plugin
    一款基于各大企业信息API的工具，解决在遇到的各种针对国内企业信息收集难题。一键收集控股公司ICP备案、APP、小程序、微信公众号等信息聚合导出。
    https://github.com/wgpsec/ENScan_GO
    """
    def __init__(self):
        self.TYPE = "infoGather"
        self.Target = []
        self.Result = {}
        self.Save = {
            "OutputPath":[],
            "Command":[],
            "CommandResult":[]
            }
        self.pluginResult = {
            "target": [],
            "result": {},
            "save": {}
        }
        
    def extractData(self, path):
        """提取一些工具输出的xlsx文件中的通用信息

        Returns:
            _type_: _description_
        """
        try:
            def extractICP():
                result = []
                ws = resultXlsx['ICP备案']
                if ws["C1"].value != "域名":
                    raise Exception("ICP备案的结果不符合预期")
                for row in ws[f'C2:C{ws.max_row}']:
                    for cell in row:
                        result.append(cell.value)
                pInfo(result)
                return result
            def extractAPP():
                result = []
                ws = resultXlsx['APP']
                if ws["A1"].value != "名称":
                    raise Exception("APP信息的结果不符合预期")
                for row in ws[f'A2:A{ws.max_row}']:
                    for cell in row:
                        result.append(cell.value)
                return result            
            import os
            fileList = os.listdir(path)
            from openpyxl import load_workbook
            for _fl in fileList:
                resultXlsx = load_workbook(
                    os.path.join(path, _fl)
                )
                self.Result["subdomain"] = extractICP()
                self.Result["app"] = extractAPP()
                continue
        except Exception as e:
            pInfo(f"|--插件提取数据出错：{e}")
            self.Result["subdomain"] = []
            self.Result["app"] = []

    def setEnv(self, kwargs):
        """
        设置插件使用代理/设定运行结果保存位置/拼接命令
        """
        pInfo(self.__doc__)
        import time

        outPath = f"out/infoGather_ENScan/{int(round(time.time() * 1000))}"
        self.Save["OutputPath"].append(outPath)
        pInfo(f"|--设定结果导出位置：{outPath}")
        myProxy = "https://" + kwargs["config"]["proxy"]["https"]
        pInfo(f"|--正在使用代理：{myProxy}")
        cmd = kwargs["config"]["cmd"]
        import os 
        if not os.path.exists(outPath):
            os.makedirs(outPath)
        cmd = cmd.format(outPath, kwargs["config"]["Target"][0])
        pInfo(f"|--设定执行命令：{cmd}")
        self.Save["Command"].append(cmd)
        return outPath, myProxy, cmd
        
    def getResult(self, cmd):
        import subprocess
        res = subprocess.Popen(cmd, shell=True, stdout=subprocess.PIPE, stderr=subprocess.PIPE)
        output, error_msgs = res.communicate()
        import base64
        self.Save["CommandResult"].append(base64.b64encode(output).decode("utf8"))
        print(output.decode(encoding='utf-8'))

    def resultFilter(self, kwargs):
        """
        安装相邻执行的两个插件的需要对target进行修改
        将Target变为ICP备案的域名列表
        """
        assert(type(kwargs["config"]["Target"]) == type([]))
        self.pluginResult["target"] = self.Result["subdomain"] + kwargs["config"]["Target"]
        self.pluginResult["target"] = list(set(self.pluginResult["target"]))
        self.pluginResult["result"] = self.Result
        self.pluginResult["save"] = self.Save
        return self.pluginResult
    
    def run(self, *args, **kwargs):
        """
        run 运行插件，Based on https://github.com/wgpsec/ENScan_GO

        Args:
            kwargs["config"]: 导入的配置
        
        Returns:
            dict: {
                target(list):下一个插件的待测目标,
                result(dict):当前此插件的经过处理的运行结果,
                save(dict):此插件运行过程中的信息存档
            }
        """
        outPath, myProxy, cmd = self.setEnv(kwargs)
        self.getResult(cmd)
        self.extractData(outPath)
        self.pluginResult = self.resultFilter(kwargs)
        return self.pluginResult